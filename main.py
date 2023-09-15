import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

from fuzzywuzzy import process
from fuzzywuzzy import fuzz

from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent
parser_files_dir = os.listdir(f"{BASE_DIR}/files_to_parse")


# https://pypi.org/project/fuzzywuzzy/
# Процент сходства 
# Если сходство болле FUZZY_PERCENT то слово отбрасывается
FUZZY_PERCENT = 59


# Слова которые необходимо вырезать из имени исполнителя
WORDS_TO_CUT = [
    'для', 'оглы',
]


class ExcelAggregateData:
    """
    Запись данных в таблицу aggregator
    1) исполнитель
    2) количество оформленных  патентов
    """

    def make_dict_data(self, data):
        data_list = []
        for obj in data:
            for i in obj:
                data_list.append(i[1])
        return data_list

    def write_new_file(self, data):
        from collections import Counter
        wb = Workbook()
        ws = wb.active
        new_data = self.make_dict_data(data)
        res = Counter(new_data).most_common()
        # add column headings. NB. these must be strings
        ws.append(["Патентообладатель", "Количество патентов"])
        for i in res:
            ws.append(i)
        tab = Table(displayName="Table1", ref="A1:B2")
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True,
                               showColumnStripes=True
                               )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        time_created = datetime.now().strftime("%d-%m-%Y_%H:%M")
        
        wb.save(f"aggregator_{time_created}.xlsx")
        del new_data


class ExcelPatentData:
    """
    Запись данных в таблицы result, deleted
    1) патентообладатель
    2) исполнитель
    3) адрес исполнителя

    """
    def write_new_file(self, data, filename, rows_count):
        #print('worksheet data', data)
        wb = Workbook()
        ws = wb.active
        # add column headings. NB. these must be strings
        ws.append(["Патентообладатель", "Исполнитель", "Адрес исполнителя"])
        for row in data:
            for i in row:
                ws.append(i)
        tab = Table(displayName="Table1", ref="A1:B2")

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True,
                               showColumnStripes=True
                               )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        time_created = datetime.now().strftime("%d-%m-%Y_%H:%M")
        # Устанавливаем правило для выделения цветом в файле deleted
        bg_red = PatternFill(fill_type='solid', fgColor="FFC7CE")
        dxf = DifferentialStyle(fill=bg_red)
        rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        rule.formula = ['$D1="deleted"']
        ws.conditional_formatting.add(f"A1:D{rows_count*2}", rule)
        # Сохраняем документ
        wb.save(f"{filename}_{time_created}.xlsx")


class ExcelParser:
    """
    Чтение данны из файла с помощью openpyxl
    # https://openpyxl.readthedocs.io/en/stable/
    Обработка текста с помощью регулярных выражений
    """
    def __init__(self):
        self.path_to_folder = f"{BASE_DIR}/files_to_parse/"

    def format_is_valid(self, file):
        *_, format = file.split(".")
        return format in ["xlsx", "xlsm", "xls", "xltx"]

    def file_path(self, file):
        return f"{self.path_to_folder}{file}"

    def parse_files(self):
        for file in parser_files_dir:
            if not self.format_is_valid(file):
                continue
            file_path = self.file_path(file)
            data, deleted_data, rows_count = self.read(file_path)
            file_manager = ExcelPatentData()
            file_manager.write_new_file(data, "result", rows_count)

            file_manager.write_new_file(deleted_data, "deleted", rows_count)
            ExcelAggregateData().write_new_file(data)

        return f"Files Parse completed see result in: `{BASE_DIR}`"

    def read(self, path_to_file):
        
        data_frame = load_workbook(path_to_file)
        frame = data_frame.active
        #rows_count = 100
        rows_count = len(frame["A"])
        rows_gen = (i for i in range(2, rows_count))
        #rows_gen = (i for i in range(2, 10))
        data, deleted_data = self.read_from_table(rows_gen, frame)
        return data, deleted_data, rows_count

    def read_from_table(self, rows, frame, data=[], deleted_data=[]):
        for row in rows:
            string_data = []
            for col in frame.iter_cols(10, 12):
                # print('row number', col[row].value)
                string_data.append(col[row].value)
            new_data, del_data = self.parse_string(string_data)
            data.append(new_data)
            deleted_data.append(del_data)
        return data, deleted_data

    def parse_string(self, string_data: list):
        """
        Парсим строку из таблицы
        """
        new_list = []
        deleted_data = []
        # Получаем список всех патентообладателей
        patent_owners = self.get_members(string_data[0])
        # print('patent_owners list for iteration', )
        # Получаем значение адреса
        address = string_data[2]
        address_upper = address.upper()
        requester = self.legal_entity_or_individual(address)
        authors_list = []
        for owner in patent_owners:

            # print('patent_owners', patent_owners)
            # Получаем патентообладателя(фио или название компании)
            author = self.author_fio_or_company_title(owner)
            if author is None:
                # print("AUTHOR is NONE", author)
                deleted_data.append((string_data[0], requester, address, "deleted"))
                continue

            author_upper = author.upper()

            if author_upper in address_upper:
                deleted_data.append((string_data[0], requester, address, "deleted"))

            if requester.upper().replace(" ", "") in author_upper.replace(" ", ""):
                deleted_data.append((string_data[0], requester, address, "deleted"))

            requester = requester.replace(".", " ")

            if author_upper.replace(".", " ") == requester.upper():
                deleted_data.append((string_data[0], requester, address, "deleted"))

            if self.number_in_string(requester):
                deleted_data.append((string_data[0], requester, address, "deleted"))

            author_cor = self.fio_corrector(author_upper)
            if author_cor in self.fio_corrector(requester.upper()):
                deleted_data.append((string_data[0], requester, address, "deleted"))

            if self.words_firs_indexs(author_upper) in requester.upper():
                deleted_data.append((string_data[0], requester, address, "deleted"))
            else:
                authors_list.append(author)

        if not self.number_in_string(requester) and len(authors_list) > 0:
            if self.fuzzy_algorim(authors_list, requester):
                deleted_data.append((string_data[0], requester, address, "deleted"))
            else:
                authors = "\n".join(i for i in authors_list)
                print("Authors res ==>", authors)
                new_list.append((authors, self.cut_exc_words(requester), address))
                deleted_data.append((string_data[0], requester, address, "Not deleted"))

        return new_list, deleted_data

    def get_members(self, string):
        """
        Возвращает имя компании
          или ФИО физ лица
        """
        if string is None:
            return []
        return [i.strip() for i in re.split(r'\n', re.sub(r'[)(A-Z]', "", string).strip())]

    def update_owner_list(self, author, patent_owners):
        owner_list = patent_owners.remove(author) if author in patent_owners else patent_owners
        if not owner_list:
            return None
        return "\n".join(i for i in owner_list)

    def cut_exc_words(self, string):
        """
        Вырезаем слова исключения и стороки и точки
        при сокращени имени и отчества
        """
        string = string
        for word in WORDS_TO_CUT:
            # print('pattern', word)
            string = string.replace(word, "")
        return string.replace(".", " ").strip()

    def author_fio_or_company_title(self, string):
        """
        Получаем фио автора или название фирмы
        """
        try:
            string = self.cut_exc_words(string)
            fio = re.split(" ", string)
            len_fio = len(fio)
            if len_fio in [3, 4]:
                # print('fio  fio', fio)
                company_name = self.get_company_name(string)
                if len(company_name) != 0:
                    return company_name[0]
                res = f"{fio[0]} {fio[1][0]} {fio[2][0]}"
            elif len(fio) == 2:
                res = f"{fio[0]} {fio[1][0]}"
            else:
                res = self.legal_entity_or_individual(string)
            return res
        except:
            print("Exception in author_fio_or_company_title", string)

    def correct_family(self, family):
        if family[-1] in ["А", "У"]:
            return family[0:-1]
        if family[-2:] in ["ОЙ"]:
            return family[0:-2]
        else:
            return family

    def fio_corrector(self, string):
        """
        Дополнительная проверка фамилии на склонения
        """
        # print('fio_corrector IN', string)
        try:
            fio_data = re.split(" ", string)
            if len(fio_data[0]) > 1:
                family = self.correct_family(fio_data[0])
                fio = f"{family} {fio_data[1][0]} {fio_data[2][0]}"
            else:
                family = self.correct_family(fio_data[2])
                fio = f"{family} {fio_data[0][0]} {fio_data[1][0]}"
            # print('fio_corrector OUT', fio)
            return fio
        except:
            print("Except fio_converter string", string)
            return string

    def get_company_name(self, string):
        """
        Ищет имя компании которое находиться в ковычках
        """
        return re.findall(r'"(.*)"', string)

    def tc_exist(self, string):
        """
        Проверка строки адреса на БЦ(торговый центр)
        """
        trading_center = re.findall(r'\БЦ', string)
        return len(trading_center) > 0

    def legal_entity_or_individual(self, string):
        """
        Из адресса получаем название фирмы в ковычках
        или ФИО автора исключаем БЦ
        """
        try:
            # Проверяем если в адресе торговый центр
            if self.tc_exist(string):
                list_data = self.get_company_name(string)
                if len(list_data) > 1:
                    return list_data[-1]
                else:
                    list_data = re.split(",",  string)
                    if self.tc_exist(string):
                        # Если нет занятой и из строки не удалился БЦ
                        return " ".join(i for i in list_data[-1].split(" ")[-2:])
                    return self.cut_exc_words(list_data[-1])

            list_data = self.get_company_name(string)
            if len(list_data) != 0:
                # Если нашлось название фирымы в ковычках 

                return list_data[0]
            fio = string
            # Допустим что фамилия имя и отчество
            # в сумме не может быть больше 35 символов
            if len(string) < 35:
                # Если  название фирмы в ковычках не нашлось
                list_data = re.split(",", string)
                if len(list_data) != 0:
                    fio = self.cut_exc_words(list_data[-1])
                    if self.number_in_string(fio):
                       fio = " ".join(i for i in (re.split(" ", fio)[-2:]))
                else:
                    # Если в адресе есть фамилия
                    fio = f"{fio[0]} {fio[1][0]} {fio[2][0]}"
                # print("fio in address string", fio)
                return fio
            else:
                # Ищем фамилию в конце строки
                fio = re.split(",", string)[-1].strip()
                # Бывают случаи когда в конце вместе с фамилией
                #  адрес или ещё какие либо слова проверяем этот случай
                if self.number_in_string(fio):
                    # print('number_in_string, fio===>>', fio)
                    fio = " ".join(i for i in (re.split(" ", fio)[-2:]))
                return fio
        except:
            raise Exception(f"Invalid data company {string}")

    def address_author_fio(self, address):
        try:
            address = (address.replace(".", " ")).strip()
            fio = re.split(" ", address)[-3:]
            return f"{fio[0]} {fio[1][0]} {fio[2][0]}"
        except:
            print("Exception addres author fio", address)

    def number_in_string(self, string):
        list_number = re.findall(r'[0-9]', string)
        return len(list_number) > 0

    def words_firs_indexs(self, string):
        res = []
        i_list = string.split(" ")
        if len(i_list) <= 1:
            return string
        for i in i_list:
            if i:
                res.append(i)
            else:
                continue
        return "".join(i[0] for i in res)
    
    def fuzzy_algorim(self, author_list, requester):
        """
        Анализ текста с помощью алгоритма Левенштейна для анализа текста
        Проверяем есть ли заказцик в патентообладателях с учётом погрешности
        https://pypi.org/project/fuzzywuzzy/
        """
        print('fuzzy_algorim data', requester, author_list)

        percent = process.extractOne(requester, author_list)[1]
        print("Word similarity percent:", percent)
        res = percent > FUZZY_PERCENT
        if not res:
            for i in author_list:
                if len(i) > 35:
                    percent = fuzz.ratio(requester, i)
                    print("Word similarity percent level 2:", percent)
                    res = percent > FUZZY_PERCENT
                    if not res:
                        percent = process.extractOne(requester, [i for i in (re.split(" ", i))])[1]
                        print("Word similarity percent level 3====>>>>>:", percent)
                        res = percent > FUZZY_PERCENT
        return res


if __name__ == "__main__":
    """
    На выходе в рабочей директории:
       - Первый документ
        отфильтрованный список с полями:
        1) патентообладатель
        2) исполнитель
        3) адрес исполнителя

        - Второй документ:
        Копия оригинального файла с выделением 
        удалённых строк

        - Третий документ:
        1) исполнитель
        2) количество оформленных  патентов
    """
    parser = ExcelParser()
    res = parser.parse_files()
    print("Parse completed:", res)
